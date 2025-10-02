# utils/safety_stock/export.py
"""
Export and reporting functions for Safety Stock Management
Updated for simplified DB structure (no min/max stock, 3 methods only)
"""

import pandas as pd
import io
from datetime import datetime
from typing import Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import text
from ..db import get_db_engine

logger = logging.getLogger(__name__)

# Excel formatting constants
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def format_excel_sheet(worksheet, freeze_row: int = 2):
    """Apply standard formatting to Excel worksheet"""
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    for row in worksheet.iter_rows(min_row=1):
        for cell in row:
            cell.border = THIN_BORDER
    
    worksheet.freeze_panes = f'A{freeze_row}'


def export_to_excel(
    df: pd.DataFrame,
    include_parameters: bool = True,
    include_metadata: bool = True
) -> io.BytesIO:
    """Export safety stock data to formatted Excel file"""
    output = io.BytesIO()
    
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main columns (removed min/max stock)
            main_columns = [
                'pt_code', 'product_name', 'brand_name',
                'entity_code', 'entity_name',
                'customer_code', 'customer_name',
                'safety_stock_qty', 'reorder_point', 'reorder_qty',
                'calculation_method', 'rule_type', 'status',
                'effective_from', 'effective_to',
                'priority_level', 'business_notes'
            ]
            
            if include_metadata:
                main_columns.extend(['created_by', 'created_date', 'updated_by', 'updated_date'])
            
            export_columns = [col for col in main_columns if col in df.columns]
            main_df = df[export_columns].copy()
            
            for col in ['effective_from', 'effective_to', 'created_date', 'updated_date']:
                if col in main_df.columns:
                    main_df[col] = pd.to_datetime(main_df[col], errors='coerce').dt.strftime('%Y-%m-%d')
            
            main_df.to_excel(writer, sheet_name='Safety Stock Levels', index=False)
            
            if include_parameters:
                param_df = _prepare_parameters_sheet(df)
                if not param_df.empty:
                    param_df.to_excel(writer, sheet_name='Calculation Parameters', index=False)
            
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                format_excel_sheet(workbook[sheet_name])
        
        output.seek(0)
        logger.info(f"Exported {len(df)} safety stock records to Excel")
        return output
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise


def _prepare_parameters_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare calculation parameters sheet"""
    param_columns = [
        'pt_code', 'product_name', 'entity_code', 'customer_code',
        'calculation_method', 'lead_time_days', 'safety_days',
        'service_level_percent', 'avg_daily_demand', 'demand_std_deviation',
        'last_calculated_date'
    ]
    
    available_columns = [col for col in param_columns if col in df.columns]
    if not available_columns:
        return pd.DataFrame()
    
    param_df = df[available_columns].copy()
    
    if 'calculation_method' in param_df.columns:
        param_df = param_df[param_df['calculation_method'].notna()]
    
    if 'last_calculated_date' in param_df.columns:
        param_df['last_calculated_date'] = pd.to_datetime(
            param_df['last_calculated_date'], errors='coerce'
        ).dt.strftime('%Y-%m-%d')
    
    return param_df


def create_upload_template(include_sample_data: bool = False) -> io.BytesIO:
    """Create Excel template for bulk upload - 3 methods only"""
    output = io.BytesIO()
    
    try:
        # Template structure (removed min/max stock)
        template_data = {
            'product_id': ['Required: Product ID'],
            'entity_id': ['Required: Entity ID'],
            'customer_id': ['Optional: Customer ID (leave blank for general rule)'],
            'safety_stock_qty': ['Required: Safety Stock Quantity'],
            'reorder_point': ['Optional: Reorder Point'],
            'reorder_qty': ['Optional: Reorder Quantity'],
            'calculation_method': ['Optional: FIXED | DAYS_OF_SUPPLY | LEAD_TIME_BASED'],
            'lead_time_days': ['Optional: Lead Time (for LEAD_TIME_BASED)'],
            'safety_days': ['Optional: Safety Days (for DAYS_OF_SUPPLY)'],
            'service_level_percent': ['Optional: Service Level % (for LEAD_TIME_BASED: 90, 95, 98, 99)'],
            'demand_std_deviation': ['Optional: Demand Std Dev (for LEAD_TIME_BASED)'],
            'avg_daily_demand': ['Optional: Average Daily Demand'],
            'effective_from': ['Required: Start Date (YYYY-MM-DD)'],
            'effective_to': ['Optional: End Date (YYYY-MM-DD)'],
            'priority_level': ['Optional: Priority (default 100)'],
            'business_notes': ['Optional: Business Notes']
        }
        
        df = pd.DataFrame(template_data)
        
        if include_sample_data:
            sample_rows = [
                {
                    'product_id': 101,
                    'entity_id': 1,
                    'customer_id': '',
                    'safety_stock_qty': 100,
                    'reorder_point': 150,
                    'reorder_qty': 200,
                    'calculation_method': 'DAYS_OF_SUPPLY',
                    'lead_time_days': '',
                    'safety_days': 14,
                    'service_level_percent': '',
                    'demand_std_deviation': '',
                    'avg_daily_demand': 10,
                    'effective_from': '2025-01-01',
                    'effective_to': '',
                    'priority_level': 100,
                    'business_notes': 'Standard rule - stable demand'
                },
                {
                    'product_id': 102,
                    'entity_id': 1,
                    'customer_id': 5,
                    'safety_stock_qty': 65,
                    'reorder_point': 150,
                    'reorder_qty': 100,
                    'calculation_method': 'LEAD_TIME_BASED',
                    'lead_time_days': 14,
                    'safety_days': '',
                    'service_level_percent': 95,
                    'demand_std_deviation': 3.5,
                    'avg_daily_demand': 8,
                    'effective_from': '2025-01-01',
                    'effective_to': '2025-12-31',
                    'priority_level': 50,
                    'business_notes': 'Customer-specific - variable demand'
                },
                {
                    'product_id': 103,
                    'entity_id': 1,
                    'customer_id': '',
                    'safety_stock_qty': 200,
                    'reorder_point': '',
                    'reorder_qty': '',
                    'calculation_method': 'FIXED',
                    'lead_time_days': '',
                    'safety_days': '',
                    'service_level_percent': '',
                    'demand_std_deviation': '',
                    'avg_daily_demand': '',
                    'effective_from': '2025-01-01',
                    'effective_to': '',
                    'priority_level': 100,
                    'business_notes': 'New product - manual input'
                }
            ]
            
            sample_df = pd.DataFrame(sample_rows)
            df = pd.concat([df, sample_df], ignore_index=True)
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Safety Stock Import', index=False)
            
            instructions_df = _create_instructions_sheet()
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            workbook = writer.book
            
            data_sheet = workbook['Safety Stock Import']
            format_excel_sheet(data_sheet)
            
            for cell in data_sheet[1]:
                cell.font = Font(italic=True, color="FF0000")
            
            inst_sheet = workbook['Instructions']
            inst_sheet.column_dimensions['A'].width = 100
            inst_sheet['A1'].font = Font(bold=True, size=14)
        
        output.seek(0)
        logger.info("Created upload template")
        return output
        
    except Exception as e:
        logger.error(f"Error creating template: {e}")
        raise


def _create_instructions_sheet() -> pd.DataFrame:
    """Create instructions dataframe for template"""
    instructions = [
        'SAFETY STOCK BULK UPLOAD TEMPLATE',
        '',
        'REQUIRED FIELDS:',
        '• product_id: Numeric ID of the product',
        '• entity_id: Numeric ID of the entity/warehouse owner',
        '• safety_stock_qty: Safety stock quantity (must be >= 0)',
        '• effective_from: Start date in YYYY-MM-DD format',
        '',
        'OPTIONAL FIELDS:',
        '• customer_id: Leave blank for general rules, provide ID for customer-specific',
        '• reorder_point/qty: Reorder trigger and quantity',
        '',
        'CALCULATION METHODS (3 options):',
        '',
        '1. FIXED - Manual input, no calculation',
        '   • Use for: New products, special contracts',
        '   • Required: None (just safety_stock_qty)',
        '',
        '2. DAYS_OF_SUPPLY - Simple coverage calculation',
        '   • Use for: Stable demand (CV < 20%)',
        '   • Required: safety_days, avg_daily_demand',
        '   • Formula: SS = safety_days × avg_daily_demand',
        '',
        '3. LEAD_TIME_BASED - Statistical method',
        '   • Use for: Variable demand (CV >= 20%), critical items',
        '   • Required: lead_time_days, service_level_percent, demand_std_deviation',
        '   • Formula: SS = Z-score × √lead_time × std_deviation',
        '   • Service levels: 90, 94, 95, 98, 99',
        '',
        'NOTES:',
        '• Priority: Lower number = higher priority (default 100)',
        '• Customer-specific rules override general rules',
        '• Delete the first row (field descriptions) before uploading',
        '• See sample data rows for examples of each method'
    ]
    
    return pd.DataFrame({'Instructions': instructions})


def generate_review_report(
    review_period_days: int = 30,
    entity_id: Optional[int] = None
) -> io.BytesIO:
    """Generate comprehensive review report (simplified)"""
    output = io.BytesIO()
    
    try:
        engine = get_db_engine()
        
        summary_df = _get_performance_summary(engine, review_period_days, entity_id)
        pending_df = _get_pending_reviews(engine, review_period_days, entity_id)
        recent_df = _get_recent_reviews(engine, review_period_days, entity_id)
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            if not pending_df.empty:
                pending_df.to_excel(writer, sheet_name='Pending Reviews', index=False)
            
            if not recent_df.empty:
                recent_df.to_excel(writer, sheet_name='Recent Reviews', index=False)
            
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                format_excel_sheet(workbook[sheet_name])
        
        output.seek(0)
        logger.info(f"Generated review report for {review_period_days} days")
        return output
        
    except Exception as e:
        logger.error(f"Error generating review report: {e}")
        raise


def _get_performance_summary(engine, days: int, entity_id: Optional[int]) -> pd.DataFrame:
    """Get performance summary statistics"""
    query = text("""
    SELECT 
        COUNT(DISTINCT s.id) as total_items,
        COUNT(DISTINCT CASE WHEN ssp.calculation_method = 'FIXED' THEN s.id END) as fixed_method,
        COUNT(DISTINCT CASE WHEN ssp.calculation_method = 'DAYS_OF_SUPPLY' THEN s.id END) as days_of_supply,
        COUNT(DISTINCT CASE WHEN ssp.calculation_method = 'LEAD_TIME_BASED' THEN s.id END) as lead_time_based,
        COUNT(DISTINCT ssr.id) as total_reviews
    FROM safety_stock_levels s
    LEFT JOIN safety_stock_parameters ssp ON s.id = ssp.safety_stock_level_id
    LEFT JOIN safety_stock_reviews ssr ON s.id = ssr.safety_stock_level_id
        AND ssr.review_date >= DATE_SUB(CURRENT_DATE(), INTERVAL :days DAY)
    WHERE s.delete_flag = 0 AND s.is_active = 1
    """ + (" AND s.entity_id = :entity_id" if entity_id else ""))
    
    params = {'days': days}
    if entity_id:
        params['entity_id'] = entity_id
    
    with engine.connect() as conn:
        result = conn.execute(query, params).fetchone()
    
    return pd.DataFrame([{
        'Metric': 'Total Active Items',
        'Value': result.total_items
    }, {
        'Metric': 'FIXED Method',
        'Value': result.fixed_method
    }, {
        'Metric': 'DAYS_OF_SUPPLY Method',
        'Value': result.days_of_supply
    }, {
        'Metric': 'LEAD_TIME_BASED Method',
        'Value': result.lead_time_based
    }, {
        'Metric': f'Reviews in Last {days} Days',
        'Value': result.total_reviews
    }])


def _get_pending_reviews(engine, days: int, entity_id: Optional[int]) -> pd.DataFrame:
    """Get items pending review"""
    query = text("""
    SELECT 
        p.pt_code,
        p.name as product_name,
        e.company_code as entity_code,
        s.safety_stock_qty,
        ssp.calculation_method,
        DATEDIFF(CURRENT_DATE(), ssp.last_calculated_date) as days_since_calc,
        MAX(ssr.review_date) as last_review_date
    FROM safety_stock_levels s
    JOIN products p ON s.product_id = p.id
    JOIN companies e ON s.entity_id = e.id
    LEFT JOIN safety_stock_parameters ssp ON s.id = ssp.safety_stock_level_id
    LEFT JOIN safety_stock_reviews ssr ON s.id = ssr.safety_stock_level_id
    WHERE s.delete_flag = 0 AND s.is_active = 1
    """ + (" AND s.entity_id = :entity_id" if entity_id else "") + """
    GROUP BY s.id, p.pt_code, p.name, e.company_code, s.safety_stock_qty, 
             ssp.calculation_method, ssp.last_calculated_date
    HAVING (
        last_review_date IS NULL 
        OR last_review_date < DATE_SUB(CURRENT_DATE(), INTERVAL :days DAY)
        OR days_since_calc > :days
    )
    ORDER BY days_since_calc DESC
    LIMIT 100
    """)
    
    params = {'days': days}
    if entity_id:
        params['entity_id'] = entity_id
    
    with engine.connect() as conn:
        return pd.read_sql(query, conn, params=params)


def _get_recent_reviews(engine, days: int, entity_id: Optional[int]) -> pd.DataFrame:
    """Get recent review history (simplified fields)"""
    query = text("""
    SELECT 
        ssr.review_date,
        p.pt_code,
        p.name as product_name,
        ssr.old_safety_stock_qty,
        ssr.new_safety_stock_qty,
        ssr.change_percentage,
        ssr.action_taken,
        ssr.action_reason,
        ssr.reviewed_by
    FROM safety_stock_reviews ssr
    JOIN safety_stock_levels s ON ssr.safety_stock_level_id = s.id
    JOIN products p ON s.product_id = p.id
    WHERE ssr.review_date >= DATE_SUB(CURRENT_DATE(), INTERVAL :days DAY)
    """ + (" AND s.entity_id = :entity_id" if entity_id else "") + """
    ORDER BY ssr.review_date DESC
    LIMIT 100
    """)
    
    params = {'days': days}
    if entity_id:
        params['entity_id'] = entity_id
    
    with engine.connect() as conn:
        return pd.read_sql(query, conn, params=params)